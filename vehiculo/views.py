from django.shortcuts import render, redirect
from django.shortcuts import get_object_or_404, redirect
from .models import Vehiculo
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import HttpResponse
from django.db.models import Q
from openpyxl import Workbook
from openpyxl.styles import Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import mm
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
from datetime import datetime


# LISTADO VEHÍCULO
def vehiculo_listado(request):
    buscar = request.GET.get("buscar", "")

    vehiculos = Vehiculo.objects.all()

    if buscar:
        vehiculos = vehiculos.filter(
            Q(patente__icontains=buscar) |
            Q(marca__icontains=buscar) |
            Q(modelo__icontains=buscar)
        )

    vehiculos = vehiculos.order_by("-carga")


    return render(request, "vehiculo_listado.html", {
        "vehiculos": vehiculos,
        "buscar": buscar,
    })





# REGISTRO VEHICULO
def registro_vehiculo(request):
    if request.method == 'POST':
        patente = request.POST.get('patente')
        marca = request.POST.get('marca')
        modelo = request.POST.get('modelo')
        tipo_vehiculo = request.POST.get('tipo_vehiculo')
        anio = request.POST.get('anio')
        n_motor = request.POST.get('n_motor')
        carga = request.POST.get('carga')
        propietario = request.POST.get('propietario')
        lugar_mantencion = request.POST.get('lugar_mantencion')

        activo = request.POST.get('activo') == 'on'

        Vehiculo.objects.create(
            patente=patente,
            marca=marca,
            modelo=modelo,
            tipo_vehiculo=tipo_vehiculo,
            anio=anio,
            n_motor=n_motor,
            carga=carga,
            propietario=propietario,
            lugar_mantencion=lugar_mantencion,
            activo=activo
        )

        return redirect('vehiculo_listado')

    return render(request, 'registro_vehiculo.html')





# EDITAR VEHICULO
def editar_vehiculo(request, id):
    vehiculo = get_object_or_404(Vehiculo, id=id)

    if request.method == 'POST':
        vehiculo.patente = request.POST.get('patente')
        vehiculo.marca = request.POST.get('marca')
        vehiculo.modelo = request.POST.get('modelo')
        vehiculo.tipo_vehiculo = request.POST.get('tipo_vehiculo')
        vehiculo.anio = request.POST.get('anio')
        vehiculo.n_motor = request.POST.get('n_motor')
        vehiculo.carga = request.POST.get('carga')
        vehiculo.propietario = request.POST.get('propietario')
        vehiculo.lugar_mantencion = request.POST.get('lugar_mantencion')
        vehiculo.activo = request.POST.get('activo') == 'on'

        vehiculo.save()

        return redirect('vehiculo_listado')

    return render(request, 'editar_vehiculo.html', {
        'vehiculo': vehiculo
    })





#CAMBIAR EL ESTADO DEL VEHICULO
def cambiar_estado_vehiculo(request, id):
    vehiculo = get_object_or_404(Vehiculo, id=id)

    vehiculo.activo = not vehiculo.activo
    vehiculo.save()

    return redirect('vehiculo_listado')  # Cambia por el nombre de tu vista de listado




# ============================================================
# REPORTE VEHÍCULOS - EXCEL
# ============================================================

def vehiculo_reporte_excel(request):

    buscar = request.GET.get("buscar", "")

    vehiculos = Vehiculo.objects.all()

    if buscar:
        vehiculos = vehiculos.filter(
            Q(patente__icontains=buscar) |
            Q(marca__icontains=buscar) |
            Q(modelo__icontains=buscar)
        )

    vehiculos = vehiculos.order_by("-carga")

    # Crear archivo Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Vehículos"

    # Título
    ws.merge_cells("A1:K1")
    ws["A1"] = "REPORTE DE VEHÍCULOS"
    ws["A1"].font = ws["A1"].font.copy(
        bold=True,
        size=16
    )
    ws["A1"].alignment = Alignment(horizontal="center")

    # Fecha
    ws.merge_cells("A2:K2")
    ws["A2"] = f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].alignment = Alignment(horizontal="center")

    # Encabezados
    encabezados = [
        "#",
        "Patente",
        "Tipo",
        "Marca",
        "Modelo",
        "Año",
        "N° Motor",
        "Carga",
        "Propietario",
        "Lugar Mantención",
        "Estado",
    ]

    for col, encabezado in enumerate(encabezados, start=1):
        celda = ws.cell(row=4, column=col)
        celda.value = encabezado
        celda.font = celda.font.copy(bold=True)
        celda.alignment = Alignment(horizontal="center")

    # Datos
    for fila, veh in enumerate(vehiculos, start=5):

        ws.cell(fila, 1).value = fila - 4
        ws.cell(fila, 2).value = veh.patente
        ws.cell(fila, 3).value = veh.get_tipo_vehiculo_display()
        ws.cell(fila, 4).value = veh.marca or ""
        ws.cell(fila, 5).value = veh.modelo or ""
        ws.cell(fila, 6).value = veh.anio or ""
        ws.cell(fila, 7).value = veh.n_motor or ""
        ws.cell(fila, 8).value = veh.carga or ""
        ws.cell(fila, 9).value = veh.propietario or ""
        ws.cell(fila, 10).value = veh.lugar_mantencion or ""
        ws.cell(fila, 11).value = "ACTIVO" if veh.activo else "INACTIVO"

    # Ancho de columnas
    anchos = {
        "A": 6,
        "B": 14,
        "C": 20,
        "D": 18,
        "E": 18,
        "F": 10,
        "G": 20,
        "H": 15,
        "I": 25,
        "J": 25,
        "K": 12,
    }

    for columna, ancho in anchos.items():
        ws.column_dimensions[columna].width = ancho

    # Respuesta
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = (
        'attachment; filename="reporte_vehiculos.xlsx"'
    )

    wb.save(response)

    return response





# ============================================================
# REPORTE VEHÍCULOS - PDF
# ============================================================

def vehiculo_reporte_pdf(request):

    buscar = request.GET.get("buscar", "")

    vehiculos = Vehiculo.objects.all()

    if buscar:
        vehiculos = vehiculos.filter(
            Q(patente__icontains=buscar) |
            Q(marca__icontains=buscar) |
            Q(modelo__icontains=buscar)
        )

    vehiculos = vehiculos.order_by("-carga")

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = (
        'attachment; filename="reporte_vehiculos.pdf"'
    )

    # Documento horizontal para que quepan todas las columnas
    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        rightMargin=8 * mm,
        leftMargin=8 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
    )

    estilos = getSampleStyleSheet()

    titulo = ParagraphStyle(
        "TituloReporte",
        parent=estilos["Title"],
        alignment=TA_CENTER,
        fontSize=16,
        spaceAfter=6,
    )

    texto_centrado = ParagraphStyle(
        "TextoCentrado",
        parent=estilos["Normal"],
        alignment=TA_CENTER,
        fontSize=8,
    )

    elementos = []

    # Título
    elementos.append(
        Paragraph("REPORTE DE VEHÍCULOS", titulo)
    )

    # Fecha
    fecha = datetime.now().strftime("%d/%m/%Y %H:%M")

    elementos.append(
        Paragraph(
            f"Fecha de generación: {fecha}",
            texto_centrado
        )
    )

    elementos.append(Spacer(1, 8))

    # Total
    elementos.append(
        Paragraph(
            f"Total de vehículos: {vehiculos.count()}",
            estilos["Normal"]
        )
    )

    elementos.append(Spacer(1, 8))

    # Encabezados
    datos = [[
        "#",
        "Patente",
        "Tipo",
        "Marca",
        "Modelo",
        "Año",
        "N° Motor",
        "Carga",
        "Propietario",
        "Lugar Mantención",
        "Estado",
    ]]

    # Datos
    for contador, veh in enumerate(vehiculos, start=1):

        datos.append([
            str(contador),
            veh.patente or "",
            veh.get_tipo_vehiculo_display() or "",
            veh.marca or "",
            veh.modelo or "",
            str(veh.anio or ""),
            veh.n_motor or "",
            veh.carga or "",
            veh.propietario or "",
            veh.lugar_mantencion or "",
            "ACTIVO" if veh.activo else "INACTIVO",
        ])

    # Tabla
    tabla = Table(
        datos,
        repeatRows=1,
        colWidths=[
            8 * mm,
            22 * mm,
            30 * mm,
            25 * mm,
            25 * mm,
            14 * mm,
            30 * mm,
            20 * mm,
            35 * mm,
            38 * mm,
            20 * mm,
        ],
    )

    tabla.setStyle(
        TableStyle([
            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.HexColor("#3c8dbc")
            ),
            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white
            ),
            (
                "FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold"
            ),
            (
                "FONTSIZE",
                (0, 0),
                (-1, -1),
                7
            ),
            (
                "ALIGN",
                (0, 0),
                (-1, -1),
                "CENTER"
            ),
            (
                "VALIGN",
                (0, 0),
                (-1, -1),
                "MIDDLE"
            ),
            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.grey
            ),
            (
                "ROWBACKGROUNDS",
                (0, 1),
                (-1, -1),
                [
                    colors.white,
                    colors.HexColor("#f2f2f2")
                ]
            ),
            (
                "BOTTOMPADDING",
                (0, 0),
                (-1, 0),
                6
            ),
            (
                "TOPPADDING",
                (0, 0),
                (-1, 0),
                6
            ),
        ])
    )

    elementos.append(tabla)

    # Generar PDF
    doc.build(elementos)

    return response