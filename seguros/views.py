from datetime import date, timedelta, datetime

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import HttpResponse

from .models import Seguro
from vehiculo.models import Vehiculo

# Excel
from openpyxl import Workbook
from openpyxl.styles import Alignment

# PDF
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
)




# LISTADO SEGURO
def seguro_listado(request):

    seguros = Seguro.objects.select_related("vehiculo")

    hoy = date.today()
    limite = hoy + timedelta(days=30)

    for seguro in seguros:
        seguro.alerta_vencimiento = (
            seguro.fecha_vencimiento <= limite
        )

    return render(
        request,
        "seguro_listado.html",
        {
            "seguros": seguros,
        },
    )









#CREAR SEGURO

def seguro(request):

    if request.method == 'POST':

        vehiculo_id = request.POST.get('vehiculo')
        banco = request.POST.get('banco')
        numero_poliza = request.POST.get('numero_poliza')
        item = request.POST.get('item')
        fecha_vencimiento = request.POST.get('fecha_vencimiento')
        archivo = request.FILES.get('archivo')

        vehiculo = get_object_or_404(Vehiculo, id=vehiculo_id)

        seguro = Seguro(
            vehiculo=vehiculo,
            banco=banco,
            numero_poliza=numero_poliza,
            item=item,
            fecha_vencimiento=fecha_vencimiento,
            archivo=archivo
        )

        seguro.save()

        messages.success(request, 'Seguro creado correctamente.')

        return redirect('seguro_listado')

    vehiculos = Vehiculo.objects.filter(activo=True).order_by('patente')

    return render(
        request,
        'seguro.html',
        {
            'vehiculos': vehiculos
        }
    )





# EDITAR SEGURO
def editar_seguro(request, id):

    seguro = get_object_or_404(Seguro, id=id)

    if request.method == 'POST':

        vehiculo_id = request.POST.get('vehiculo')
        banco = request.POST.get('banco')
        numero_poliza = request.POST.get('numero_poliza')
        item = request.POST.get('item')
        fecha_vencimiento = request.POST.get('fecha_vencimiento')
        archivo = request.FILES.get('archivo')

        vehiculo = get_object_or_404(Vehiculo, id=vehiculo_id)

        seguro.vehiculo = vehiculo
        seguro.banco = banco
        seguro.numero_poliza = numero_poliza
        seguro.item = item
        seguro.fecha_vencimiento = fecha_vencimiento

        # Solo cambia el archivo si se seleccionó uno nuevo
        if archivo:
            seguro.archivo = archivo

        seguro.save()

        messages.success(request, 'Seguro actualizado correctamente.')

        return redirect('seguro_listado')

    vehiculos = Vehiculo.objects.filter(
        activo=True
    ).order_by('patente')

    return render(
        request,
        'editar_seguro.html',
        {
            'seguro': seguro,
            'vehiculos': vehiculos
        }
    )




# ELIMINAR SEGURO

def eliminar_seguro(request, id):

    seguro = get_object_or_404(Seguro, id=id)

    seguro.delete()

    messages.success(request, 'Seguro eliminado correctamente.')

    return redirect('seguro_listado')



# ============================================================
# REPORTE SEGUROS - EXCEL
# ============================================================

def seguro_reporte_excel(request):

    seguros = Seguro.objects.select_related("vehiculo")

    # Crear archivo Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Seguros"

    # Título
    ws.merge_cells("A1:K1")
    ws["A1"] = "REPORTE DE SEGUROS"
    ws["A1"].font = ws["A1"].font.copy(
        bold=True,
        size=16
    )
    ws["A1"].alignment = Alignment(horizontal="center")

    # Fecha
    ws.merge_cells("A2:K2")
    ws["A2"] = (
        f"Fecha de generación: "
        f"{datetime.now().strftime('%d/%m/%Y %H:%M')}"
    )
    ws["A2"].alignment = Alignment(horizontal="center")

    # Total
    ws.merge_cells("A3:K3")
    ws["A3"] = f"Total de seguros: {seguros.count()}"
    ws["A3"].alignment = Alignment(horizontal="center")

    # Encabezados
    encabezados = [
        "#",
        "Banco",
        "Póliza",
        "Items",
        "Tipo",
        "Marca",
        "Modelo",
        "Año",
        "Patente",
        "N° Motor",
        "Vence",
    ]

    for col, encabezado in enumerate(encabezados, start=1):
        celda = ws.cell(row=5, column=col)
        celda.value = encabezado
        celda.font = celda.font.copy(bold=True)
        celda.alignment = Alignment(horizontal="center")

    # Datos
    for fila, seguro in enumerate(seguros, start=6):

        ws.cell(fila, 1).value = fila - 5
        ws.cell(fila, 2).value = seguro.banco or ""
        ws.cell(fila, 3).value = seguro.numero_poliza or ""
        ws.cell(fila, 4).value = seguro.item or ""
        ws.cell(fila, 5).value = (
            seguro.vehiculo.tipo_vehiculo
            if seguro.vehiculo else ""
        )
        ws.cell(fila, 6).value = (
            seguro.vehiculo.marca
            if seguro.vehiculo else ""
        )
        ws.cell(fila, 7).value = (
            seguro.vehiculo.modelo
            if seguro.vehiculo else ""
        )
        ws.cell(fila, 8).value = (
            seguro.vehiculo.anio
            if seguro.vehiculo else ""
        )
        ws.cell(fila, 9).value = (
            seguro.vehiculo.patente
            if seguro.vehiculo else ""
        )
        ws.cell(fila, 10).value = (
            seguro.vehiculo.n_motor
            if seguro.vehiculo else ""
        )

        if seguro.fecha_vencimiento:
            ws.cell(fila, 11).value = (
                seguro.fecha_vencimiento.strftime("%d/%m/%Y")
            )
        else:
            ws.cell(fila, 11).value = ""

    # Ancho de columnas
    anchos = {
        "A": 6,
        "B": 18,
        "C": 18,
        "D": 15,
        "E": 20,
        "F": 18,
        "G": 18,
        "H": 10,
        "I": 14,
        "J": 20,
        "K": 15,
    }

    for columna, ancho in anchos.items():
        ws.column_dimensions[columna].width = ancho

    # Respuesta
    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

    response["Content-Disposition"] = (
        'attachment; filename="reporte_seguros.xlsx"'
    )

    wb.save(response)

    return response



# ============================================================
# REPORTE SEGUROS - PDF
# ============================================================

def seguro_reporte_pdf(request):

    seguros = Seguro.objects.select_related("vehiculo")

    response = HttpResponse(content_type="application/pdf")

    response["Content-Disposition"] = (
        'attachment; filename="reporte_seguros.pdf"'
    )

    # Documento horizontal
    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        rightMargin=8 * mm,
        leftMargin=8 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
    )

    estilos = getSampleStyleSheet()

    titulo = ParagraphStyle(
        "TituloReporteSeguro",
        parent=estilos["Title"],
        alignment=TA_CENTER,
        fontSize=16,
        spaceAfter=6,
    )

    texto_centrado = ParagraphStyle(
        "TextoCentradoSeguro",
        parent=estilos["Normal"],
        alignment=TA_CENTER,
        fontSize=8,
    )

    elementos = []

    # Título
    elementos.append(
        Paragraph(
            "REPORTE DE SEGUROS",
            titulo
        )
    )

    # Fecha
    fecha = datetime.now().strftime("%d/%m/%Y %H:%M")

    elementos.append(
        Paragraph(
            f"Fecha de generación: {fecha}",
            texto_centrado
        )
    )

    elementos.append(Spacer(1, 8))

    # Total
    elementos.append(
        Paragraph(
            f"Total de seguros: {seguros.count()}",
            estilos["Normal"]
        )
    )

    elementos.append(Spacer(1, 8))

    # Encabezados
    datos = [[
        "#",
        "Banco",
        "Póliza",
        "Items",
        "Tipo",
        "Marca",
        "Modelo",
        "Año",
        "Patente",
        "N° Motor",
        "Vence",
    ]]

    # Datos
    for contador, seguro in enumerate(seguros, start=1):

        datos.append([
            str(contador),
            seguro.banco or "",
            seguro.numero_poliza or "",
            seguro.item or "",

            (
                seguro.vehiculo.tipo_vehiculo
                if seguro.vehiculo else ""
            ),

            (
                seguro.vehiculo.marca
                if seguro.vehiculo else ""
            ),

            (
                seguro.vehiculo.modelo
                if seguro.vehiculo else ""
            ),

            (
                str(seguro.vehiculo.anio or "")
                if seguro.vehiculo else ""
            ),

            (
                seguro.vehiculo.patente
                if seguro.vehiculo else ""
            ),

            (
                seguro.vehiculo.n_motor
                if seguro.vehiculo else ""
            ),

            (
                seguro.fecha_vencimiento.strftime("%d/%m/%Y")
                if seguro.fecha_vencimiento else ""
            ),
        ])

    # Tabla
    tabla = Table(
        datos,
        repeatRows=1,
        colWidths=[
            8 * mm,    # #
            25 * mm,   # Banco
            28 * mm,   # Póliza
            20 * mm,   # Items
            30 * mm,   # Tipo
            25 * mm,   # Marca
            25 * mm,   # Modelo
            14 * mm,   # Año
            22 * mm,   # Patente
            30 * mm,   # Motor
            25 * mm,   # Vence
        ],
    )

    tabla.setStyle(
        TableStyle([
            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.HexColor("#3c8dbc")
            ),
            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white
            ),
            (
                "FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold"
            ),
            (
                "FONTSIZE",
                (0, 0),
                (-1, -1),
                7
            ),
            (
                "ALIGN",
                (0, 0),
                (-1, -1),
                "CENTER"
            ),
            (
                "VALIGN",
                (0, 0),
                (-1, -1),
                "MIDDLE"
            ),
            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.grey
            ),
            (
                "ROWBACKGROUNDS",
                (0, 1),
                (-1, -1),
                [
                    colors.white,
                    colors.HexColor("#f2f2f2")
                ]
            ),
            (
                "BOTTOMPADDING",
                (0, 0),
                (-1, 0),
                6
            ),
            (
                "TOPPADDING",
                (0, 0),
                (-1, 0),
                6
            ),
        ])
    )

    elementos.append(tabla)

    # Generar PDF
    doc.build(elementos)

    return response